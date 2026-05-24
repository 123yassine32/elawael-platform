# -*- coding: utf-8 -*-
"""Analyse approfondie : bénéficiaires, dates, marques."""
from openpyxl import load_workbook
import os, sys, json, re

sys.stdout.reconfigure(encoding='utf-8')

DIR = r"C:\Users\PC\Desktop\الغياب اليومي لجميع الشهور"
files = {
    'fevrier' : ("الغياب اليومي لشهر فبراير (1).xlsx", '02'),
    'mars'    : ("الغياب اليومي لشهر مارس (1).xlsx", '03'),
    'avril'   : ("الغياب اليومي لشهر ابريل (1) (1).xlsx", '04'),
    'mai'     : ("الغياب اليومي لشهر ماي).xlsx", '05'),
}

for label, (fname, mm) in files.items():
    fpath = os.path.join(DIR, fname)
    print(f"\n{'='*70}")
    print(f"=== {label.upper()} ({mm}/2026) — {fname}")
    print('='*70)
    wb = load_workbook(fpath, data_only=True)
    ws = wb.active

    # Headers
    print(f"  Dim: {ws.max_row}L × {ws.max_column}C")
    print("\n  --- HEADER (ligne 6 = jour semaine, ligne 7 = jour mois) ---")
    header_days = []
    for c in range(1, ws.max_column + 1):
        wkday = ws.cell(row=6, column=c).value
        daynum = ws.cell(row=7, column=c).value
        header_days.append((c, wkday, daynum))

    # Print just non-empty
    for c, wk, dn in header_days:
        if wk or dn:
            print(f"    Col {c}: jour='{wk}' n°={dn}")

    # Bénéficiaires (à partir de ligne 8)
    print("\n  --- BÉNÉFICIAIRES (lignes 8+) ---")
    benef_count = 0
    for r in range(8, ws.max_row + 1):
        # Trouver le nom (généralement dernière colonne ou première colonne à droite)
        name = ws.cell(row=r, column=ws.max_column).value
        if not name:
            # Essayer cellule 2 ou autres
            for cc in [ws.max_column, ws.max_column-1, 2, 1]:
                v = ws.cell(row=r, column=cc).value
                if v and str(v).strip() and not str(v).strip().isdigit():
                    name = v
                    break
        if name:
            benef_count += 1
            if benef_count <= 5:
                # Afficher cette ligne en entier
                vals = []
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    vals.append(str(v)[:8] if v is not None else '.')
                print(f"  R{r} '{str(name)[:30]}': {' '.join(vals)}")

    print(f"\n  TOTAL bénéficiaires: {benef_count}")

    # Stats des marques (collecter toutes les valeurs dans la zone bénéf)
    marks = {}
    for r in range(8, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                s = str(v).strip()
                if s and not s.isdigit() and len(s) <= 5:
                    marks[s] = marks.get(s, 0) + 1
    print(f"\n  --- MARQUES TROUVÉES (val: count) ---")
    for k, v in sorted(marks.items(), key=lambda x: -x[1])[:15]:
        print(f"    '{k}': {v}")

    wb.close()
