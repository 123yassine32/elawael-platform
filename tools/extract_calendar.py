# -*- coding: utf-8 -*-
"""Extrait le calendrier S2 réel (jours ouvrés vs vacances) depuis les fichiers Excel."""
from openpyxl import load_workbook
import os, sys, json

sys.stdout.reconfigure(encoding='utf-8')

DIR = r"C:\Users\PC\Desktop\الغياب اليومي لجميع الشهور"
files = [
    ("02", "الغياب اليومي لشهر فبراير (1).xlsx"),
    ("03", "الغياب اليومي لشهر مارس (1).xlsx"),
    ("04", "الغياب اليومي لشهر ابريل (1) (1).xlsx"),
    ("05", "الغياب اليومي لشهر ماي).xlsx"),
]

calendar_jours_ouvres = {}  # 'YYYY-MM-DD' → présent dans fichier (au moins 1 marque)

for mm, fname in files:
    fpath = os.path.join(DIR, fname)
    print(f"\n=== Mois {mm} : {fname} ===")
    wb = load_workbook(fpath, data_only=True)
    ws = wb.active

    # Trouver les colonnes avec un jour mois (ligne 7) — c'est un nombre 1-31
    day_cols = []  # (col, daynum)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=7, column=c).value
        if isinstance(v, (int, float)) and 1 <= v <= 31:
            day_cols.append((c, int(v)))
    print(f"  Colonnes-jours détectées: {len(day_cols)} → {[d for _,d in day_cols]}")

    # Pour chaque colonne-jour, regarder si au moins une cellule benef (rows 8+) a une marque
    jours_avec_marques = []
    for (col, daynum) in day_cols:
        count_marks = 0
        for r in range(8, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                s = str(v).strip()
                if s in ('P', 'A') or (s and len(s) <= 10):
                    count_marks += 1
        if count_marks > 0:
            jours_avec_marques.append(daynum)
            calendar_jours_ouvres[f"2026-{mm}-{daynum:02d}"] = count_marks

    print(f"  Jours avec données : {sorted(jours_avec_marques)}")

    # Détecter les vacances : jours dans day_cols mais sans marques
    jours_sans = [d for _,d in day_cols if d not in jours_avec_marques]
    if jours_sans:
        print(f"  Jours du calendrier MAIS sans données (vacances ?): {sorted(jours_sans)}")
    wb.close()

# Synthèse
print(f"\n\n{'='*60}")
print(f"CALENDRIER S2 RECONSTITUÉ — {len(calendar_jours_ouvres)} jours ouvrés")
print('='*60)
import collections
par_mois = collections.defaultdict(list)
for d in sorted(calendar_jours_ouvres.keys()):
    par_mois[d[:7]].append(d[8:10])
for mm, jours in sorted(par_mois.items()):
    print(f"  {mm}: {len(jours)} jours → {','.join(jours)}")

# Sauvegarder en JSON pour usage ultérieur
out = {
    "jours_ouvres": sorted(calendar_jours_ouvres.keys()),
    "stats_par_mois": {k: len(v) for k, v in par_mois.items()},
    "source": "Excel غياب اليومي fevrier-mai 2026"
}
with open(r"C:\Users\PC\Desktop\elawael-platform\tools\calendar_s2.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"\n✅ Sauvegardé : tools/calendar_s2.json")
